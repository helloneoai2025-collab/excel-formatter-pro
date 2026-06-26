import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Alignment
from datetime import datetime
from io import BytesIO
import zipfile
import os
from copy import copy

# 1. ตั้งค่าหน้าตาของแอป
st.set_page_config(
    page_title="Excel Formatter F1 Pro (Fixed)",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 2. Custom CSS เพื่อความสวยงามแบบ TTT Style
st.markdown("""
    <style>
        .title-main {
            background: linear-gradient(135deg, #1e3a8a 0%, #3b82f6 100%);
            color: white;
            padding: 25px;
            border-radius: 12px;
            margin-bottom: 25px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        }
        .success-box {
            background: #dcfce7;
            color: #166534;
            padding: 15px;
            border-radius: 8px;
            border-left: 5px solid #22c55e;
            margin: 10px 0;
        }
        .error-box {
            background: #fee2e2;
            color: #991b1b;
            padding: 15px;
            border-radius: 8px;
            border-left: 5px solid #ef4444;
            margin: 10px 0;
        }
        .stButton>button {
            border-radius: 8px;
            height: 3em;
            font-weight: bold;
        }
    </style>
""", unsafe_allow_html=True)

st.markdown("""
    <div class="title-main">
        <h1>📊 Excel Formatter F1 Pro</h1>
        <p>เวอร์ชันแก้ไข: รองรับข้อมูลไม่จำกัดแถว และลบแถว Total ที่ผิดพลาดออกแล้ว</p>
    </div>
""", unsafe_allow_html=True)

# 3. ฟังก์ชันช่วยคัดลอกรูปแบบเซลล์
def copy_cell_style(source_cell, target_cell):
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.alignment = copy(source_cell.alignment)

# 4. ฟังก์ชันดึงข้อมูลจากไฟล์ Excel (Logic F1)
def extract_color_data_f1(file_bytes):
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    
    # ดึงเลข PO จาก H5
    po_number = ws['H5'].value if ws['H5'].value else 'UNKNOWN'
    colors = []
    
    # เริ่มดึงข้อมูลตั้งแต่แถวที่ 20 เป็นต้นไป
    for row_idx in range(20, ws.max_row + 1):
        cell_a = ws.cell(row=row_idx, column=1)
        
        # ตรวจสอบว่าเป็นข้อมูลที่ต้องการหรือไม่ (เช็ค Blue Zone หรือรูปแบบรหัส)
        is_valid = False
        if cell_a.fill and cell_a.fill.start_color and hasattr(cell_a.fill.start_color, 'rgb'):
            if cell_a.fill.start_color.rgb == 'FF00B0F0': # Blue Zone
                is_valid = True
        
        if not is_valid and cell_a.value and isinstance(cell_a.value, str) and '/' in cell_a.value:
            is_valid = True

        if is_valid and cell_a.value and isinstance(cell_a.value, str) and '/' in cell_a.value:
            cell_j = ws.cell(row=row_idx, column=10) # Qty อยู่คอลัมน์ J
            parts = cell_a.value.split('/')
            
            if len(parts) == 2:
                code11 = parts[0].strip()
                code10 = parts[1].strip()
                qty = cell_j.value if cell_j.value else 0
                
                try:
                    qty = int(qty)
                    if qty > 0:
                        colors.append({
                            'code11': code11,
                            'code10': code10,
                            'qty': qty
                        })
                except: pass
    
    return {'po_number': po_number, 'colors': colors}

# 5. ฟังก์ชันเขียนข้อมูลลง Master Form (ตัดปัญหาแถว 41)
def process_master_form_f1(master_file_bytes, data_info):
    wb = load_workbook(BytesIO(master_file_bytes))
    ws = wb['Factory code label']
    
    # เขียนข้อมูลส่วนหัว
    ws['F5'].value = data_info['po_number']
    ws['F7'].value = datetime.now().strftime('%d/%m/%Y')
    ws['B17'].value = 'Tear-Away-Factory-ID-Label'
    
    colors = data_info['colors']
    template_row = 21 # ใช้แถว 21 เป็นต้นแบบ Format
    
    # วนลูปเขียนข้อมูลตามจำนวนที่มีจริง
    for idx, color_data in enumerate(colors):
        current_row = template_row + idx
        
        # คอลัมน์ที่ต้องเขียน
        # B: OPTION 1, C: Code10, E: Code11, F: Qty
        data_map = {
            2: 'OPTION 1',
            3: color_data['code10'],
            5: color_data['code11'],
            6: color_data['qty']
        }
        
        for col_idx, val in data_map.items():
            target = ws.cell(row=current_row, column=col_idx)
            target.value = val
            # คัดลอก Format จากแถวที่ 21 เสมอเพื่อให้เส้นขอบและฟอนต์เหมือนกัน
            copy_cell_style(ws.cell(row=template_row, column=col_idx), target)
    
    # --- [แก้ไข] ลบ Logic ที่บังคับเขียนแถว 41 ออกทั้งหมดแล้ว ---
    # ข้อมูลจะไหลลงไปเรื่อยๆ ตามจำนวนจริง ไม่โดนคำว่า Total ทับอีกต่อไป

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# 6. ส่วนการแสดงผลบนหน้าจอ (UI)
st.sidebar.markdown("### 🛠️ การตั้งค่าประมวลผล")
st.sidebar.info("แอปจะดึงข้อมูลจากไฟล์ Excel และใส่ลงใน Master Form โดยรักษาเส้นขอบเดิมไว้ทั้งหมด")

col1, col2 = st.columns(2)
with col1:
    master_file = st.file_uploader("📂 1. อัพโหลด Master Form (.xlsx)", type=['xlsx'])
with col2:
    data_files = st.file_uploader("📄 2. อัพโหลดไฟล์ข้อมูล (เลือกได้หลายไฟล์)", type=['xlsx'], accept_multiple_files=True)

if st.button("🚀 เริ่มประมวลผลข้อมูล", use_container_width=True):
    if not master_file or not data_files:
        st.markdown("<div class='error-box'>⚠️ กรุณาอัพโหลดไฟล์ให้ครบถ้วนก่อนเริ่มงาน</div>", unsafe_allow_html=True)
    else:
        results = []
        progress_bar = st.progress(0)
        master_bytes = master_file.getvalue()
        
        for i, f in enumerate(data_files):
            try:
                # สกัดข้อมูล
                info = extract_color_data_f1(f.getvalue())
                # ประมวลผลลงฟอร์ม
                final_file = process_master_form_f1(master_bytes, info)
                
                results.append({
                    'name': f"processed_{info['po_number']}_{f.name}",
                    'data': final_file.getvalue(),
                    'po': info['po_number'],
                    'count': len(info['colors'])
                })
            except Exception as e:
                st.error(f"ไฟล์ {f.name} มีปัญหา: {str(e)}")
            
            progress_bar.progress((i + 1) / len(data_files))
        
        if results:
            st.markdown(f"<div class='success-box'>✅ ประมวลผลสำเร็จ {len(results)} ไฟล์!</div>", unsafe_allow_html=True)
            
            # ปุ่มดาวน์โหลด
            d_col1, d_col2 = st.columns(2)
            with d_col1:
                st.subheader("📥 ดาวน์โหลดรายไฟล์")
                for res in results:
                    st.download_button(
                        label=f"📄 {res['name'][:40]}...",
                        data=res['data'],
                        file_name=res['name'],
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
            
            with d_col2:
                if len(results) > 1:
                    st.subheader("📦 ดาวน์โหลดแบบมัดรวม")
                    zip_buffer = BytesIO()
                    with zipfile.ZipFile(zip_buffer, 'w') as zf:
                        for res in results:
                            zf.writestr(res['name'], res['data'])
                    st.download_button(
                        label="📥 ดาวน์โหลด ZIP ทั้งหมด",
                        data=zip_buffer.getvalue(),
                        file_name="TTT_Processed_Files.zip",
                        mime="application/zip",
                        use_container_width=True
                    )
            
            # ตารางสรุป
            st.subheader("📊 ตารางสรุปรายการ")
            summary_df = pd.DataFrame(results)[['po', 'count', 'name']]
            summary_df.columns = ['เลข PO', 'จำนวน Color', 'ชื่อไฟล์ผลลัพธ์']
            st.dataframe(summary_df, use_container_width=True, hide_index=True)
