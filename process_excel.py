from openpyxl import load_workbook
from datetime import datetime
from copy import copy

def extract_color_data(file_path):
    """
    ดึงข้อมูล Color จากไฟล์ข้อมูล
    - ดึงเฉพาะแถวที่มีข้อมูล "/" ใน Column A (Color Code)
    - ข้ามแถวว่างและข้อมูลที่ไม่เป็น Color Code
    """
    wb = load_workbook(file_path)
    ws = wb.active
    
    # ดึง PO# จาก H5
    po_number = ws['H5'].value if ws['H5'].value else 'UNKNOWN'
    colors = []
    
    # ดึง Color Data จากโซนสีฟ้า (เริ่มจากแถว 20 ขึ้นไป)
    for row_idx in range(20, ws.max_row + 1):
        cell_a = ws.cell(row=row_idx, column=1)
        cell_j = ws.cell(row=row_idx, column=10)
        
        # ตรวจสอบว่า Column A มีข้อมูลและมี "/" ไหม
        if cell_a.value and isinstance(cell_a.value, str):
            # ตรวจสอบว่าเป็น Color Code (มีเครื่องหมาย /)
            if '/' in cell_a.value:
                # แยก Code
                parts = cell_a.value.split('/')
                if len(parts) == 2:
                    code11 = parts[0].strip()  # หน้า /
                    code10 = parts[1].strip()  # หลัง /
                    qty = cell_j.value if cell_j.value else 0
                    
                    # เก็บข้อมูล
                    colors.append({
                        'color_code': cell_a.value,
                        'code11': code11,
                        'code10': code10,
                        'qty': int(qty) if qty else 0
                    })
    
    return {
        'po_number': po_number,
        'colors': colors
    }

def process_master_form(master_file_path, data_info, output_path):
    """
    ประมวลผล Master Form พร้อมข้อมูล Color
    คงรูปแบบเดิมไว้ 100%
    """
    wb = load_workbook(master_file_path)
    ws = wb['Factory code label']
    
    # เติม PO# ใน F5
    ws['F5'].value = data_info['po_number']
    
    # เติม DATE ใน F7 (วันที่ปัจจุบัน)
    today = datetime.now().strftime('%d/%m/%Y')
    ws['F7'].value = today
    
    # เติม CUSTOMER ITEM CODE ใน B17
    ws['B17'].value = 'Tear-Away-Factory-ID-Label'
    
    # เติม Color Data
    colors = data_info['colors']
    
    for idx, color_data in enumerate(colors):
        row = 21 + idx
        
        # OPTION 1 ใน B
        ws[f'B{row}'].value = 'OPTION 1'
        
        # Code 10 digits (หลัง /) ใน C
        ws[f'C{row}'].value = color_data['code10']
        
        # Code 11 digits (หน้า /) ใน E
        ws[f'E{row}'].value = color_data['code11']
        
        # Quantity ใน F
        ws[f'F{row}'].value = color_data['qty']
    
    # บันทึกไฟล์
    wb.save(output_path)
    return True

def process_files(master_path, data_paths):
    """ประมวลผลหลายไฟล์พร้อมกัน"""
    results = []
    
    for data_path in data_paths:
        try:
            # ดึงข้อมูล
            data_info = extract_color_data(data_path)
            
            # ตั้งชื่อ Output File
            po_num = data_info['po_number']
            output_file = f"processed_{po_num}.xlsx"
            
            # ประมวลผล
            process_master_form(master_path, data_info, f"agent/{output_file}")
            
            results.append({
                'success': True,
                'input_file': data_path,
                'output_file': output_file,
                'po_number': po_num,
                'colors_count': len(data_info['colors'])
            })
        except Exception as e:
            results.append({
                'success': False,
                'input_file': data_path,
                'error': str(e)
            })
    
    return results
